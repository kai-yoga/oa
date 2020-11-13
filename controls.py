import requests
import time
import re
from lxml import etree
import os
import vars
import openpyxl as opl

def savePath(title):
    '''指定文件的存储路径'''
    #"广大[2019]108号（广州大学关于公布2019年6月转专业学生名单的通知）"
    pass
    perffix=title.split('[')[0] if title.split('[')[0] else '----'
    suffixPatt=re.compile('.+\[(\d+)\].+')
    year=suffixPatt.findall(title)[0] if len(suffixPatt.findall(title))>0 else '####'
    path=os.path.join(vars.filesSavePath, year + '\\' + perffix)
    if not os.path.exists(path):
        os.makedirs(path)
    return path

def setCookies(cookies):
    '''返回字典型的cookies'''
    return {item.split('=')[0]: item.split('=')[1] for item in cookies.split(";")}

def isGet(L,keyWord):
    '''获取是否已下载'''
    exists=0
    if keyWord in L:
        exists=1
    return exists

def paraseHtml(etreeHtml,headers,cookies,Z):
    '''处理一页，获取一页上的相关信息'''
    L=[]
    divs = etreeHtml.xpath("//div[@class='contentDiv']/table[contains(@class,'contentList contentList2')]//tr")
    for div in divs:
        item={}
        xpath = './/td[4]/text()'
        item["titleTime"]=div.xpath(xpath)[0] if len(div.xpath(xpath))>0 else '' ###发文时间
        xpath = './/td[2]/@title'
        item["title"] = div.xpath(xpath)[0] if len(div.xpath(xpath)) > 0 else ''  ###文章标题
        xpath = './/td[2]/a/@href'
        item["titleBh"] = div.xpath(xpath)[0] if len(div.xpath(xpath)) > 0 else ''  ###文章链接编号
        xpath='.//td[3]/text()'
        item["titleBM"]=div.xpath(xpath)[0] if len(div.xpath(xpath))>0 else '' ###发文部门
        xpath='.//td[5]/a/text()'
        item['counts']=div.xpath(xpath)[0] if len(div.xpath(xpath))>0 else '' ###阅读量
        if item["titleTime"] !='' and not isGet(Z,item["titleTime"]+item["title"]+item["titleBM"]):
            item["down"] = getFile(item["titleBh"],headers,cookies)  ###下载附件
            t=(item["titleTime"],item["title"],item["titleBh"],item["titleBM"],item["counts"],item["down"])
            L.append(t)
        else:
            pass
    return L

def getFile(titleBh,headers,cookies):
    patt=re.compile('.+\((\d+)\)')
    res=patt.findall(titleBh)[0]
    ###获取文章fileid
    url='http://oa.gzhu.edu.cn/cms/frontContent.do?method=toAttachList'
    data={
        "contentId":res
    }
    r=requests.post(url=url,data=data,headers=headers,cookies=cookies)
    time.sleep(1)
    # print(r.url)
    assert r.status_code==200
    # print(r.text)
    text=r.text
    fileIdPatt=re.compile(r'.*fileId="(\d+)"')
    fileids=fileIdPatt.findall(text)####文件号
    titlePatt=re.compile(r'.*<span.*>(.*)</span>')
    # filename='无附件'
    try:
        title=titlePatt.findall(text)[0]###文件名
        for fileId in fileids:###如果有多个附件文件
            filename=os.path.join(savePath(title),title)
            url='http://oa.gzhu.edu.cn/cms/file.do'
            params={
                "method":"toFilePreview",
                "fileId":fileId
            }
            try:
                r=requests.get(url=url,params=params,headers=headers,cookies=cookies)
            except:
                print(fileId)
            with open(filename,'wb') as f:
                f.write(r.content)
                f.close()
    except:
        filename='无附件'
    return filename


def main(z):
    '''主函数,返回已经下载的相关文件'''
    # start=end=1
    res=[]
    headers= vars.headers
    urls= vars.oaurl
    cookies=setCookies(vars.cookies)
    zdz=1
    for url in urls:
        start = end = 1
        while start<=int(end):
            r = requests.get(url=url.format(str(start)), headers=headers, cookies=cookies)
            assert r.status_code == 200
            etreeHtml=etree.HTML(r.content)
            if start==1:
                pageCountPatt = re.compile('.*共(\d+)页.*')
                end=pageCountPatt.findall(r.text)[0]
                if zdz=='0':
                    end='3'
            print('当前处理第{}页,url={}'.format(str(start),r.url)+'共{}页'.format(end))
            res.extend(paraseHtml(etreeHtml,headers,cookies,z))
            start=start+1
    return res

def initXlsx():
    L=[]
    if not os.path.exists(os.path.join(r"D:\projects\oa",'文件汇总.xlsx')):
        wb=opl.Workbook()
        ws=wb.active
        ws.append(['时间','标题','编号','部门','阅读量','文件保存路径'])
        wb.save(os.path.join(r"D:\projects\oa",'文件汇总.xlsx'))
    wb=opl.load_workbook(os.path.join(r"D:\projects\oa",'文件汇总.xlsx'))
    ws=wb.active
    for row in ws.iter_rows(min_row=2,max_row=ws.max_row,min_col=1,max_col=ws.max_column):
        sj=str(row[0].value) if row[0].value else ''
        bt=str(row[1].value) if row[1].value else ''
        bm=str(row[3].value) if row[3].value else ''
        t=sj+bt+bm
        L.append(t)
    wb.save(os.path.join(r"D:\projects\oa",'文件汇总.xlsx'))
    return L

def writeXls(res):
    '''写入xlsx'''
    wb=opl.load_workbook(os.path.join(r"D:\projects\oa",'文件汇总.xlsx'))
    ws=wb.active
    for ele in res:
        ws.append(ele)
    wb.save(os.path.join(r"D:\projects\oa", '文件汇总.xlsx'))
    return 1

if __name__=="__main__":
    L=initXlsx()
    res=main(L)
    writeXls(res)